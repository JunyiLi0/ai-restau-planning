from pathlib import Path
from datetime import datetime, timedelta
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.models.schemas import WeekPlanning, EmployeeWeekSchedule, DaySchedule, ShiftData


DAYS = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
DAY_LABELS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


def get_next_week_info() -> tuple[int, int, str, str]:
    """Retourne le numéro de semaine, l'année, et les dates de la semaine prochaine."""
    today = datetime.now()
    days_until_monday = (7 - today.weekday()) % 7
    if days_until_monday == 0:
        days_until_monday = 7
    next_monday = today + timedelta(days=days_until_monday)
    next_sunday = next_monday + timedelta(days=6)

    week_number = next_monday.isocalendar()[1]
    year = next_monday.year

    start_date = next_monday.strftime("%d/%m/%Y")
    end_date = next_sunday.strftime("%d/%m/%Y")

    return week_number, year, start_date, end_date


def get_week_dates(week_number: int, year: int) -> tuple[str, str]:
    """Calcule les dates de début et fin pour une semaine donnée."""
    jan1 = datetime(year, 1, 1)
    days_to_monday = (7 - jan1.weekday()) % 7
    if jan1.weekday() <= 3:
        first_monday = jan1 - timedelta(days=jan1.weekday())
    else:
        first_monday = jan1 + timedelta(days=days_to_monday)

    target_monday = first_monday + timedelta(weeks=week_number - 1)
    target_sunday = target_monday + timedelta(days=6)

    return target_monday.strftime("%d/%m/%Y"), target_sunday.strftime("%d/%m/%Y")


class ExcelHandler:
    def __init__(self):
        self.header_font = Font(bold=True, size=10)
        self.title_font = Font(bold=True, size=12)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=10, color="FFFFFF")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def create_planning_workbook(self, planning: WeekPlanning) -> Workbook:
        wb = Workbook()
        ws = wb.active

        # Calculer les dates
        if planning.week_number and planning.year:
            week_num = planning.week_number
            year = planning.year
            start_date, end_date = get_week_dates(week_num, year)
        else:
            week_num, year, start_date, end_date = get_next_week_info()

        # Titre de l'onglet
        ws.title = f"Semaine {week_num}"

        # Titre du document (ligne 1)
        title = f"Planning des employés WOK10 - Semaine {week_num} du {start_date} au {end_date}"
        # 1 + 7*6 + 2 = 45 colonnes
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=45)
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ligne vide
        ws.row_dimensions[2].height = 10

        self._setup_headers(ws, start_row=3)
        self._populate_data(ws, planning, start_row=5)
        self._adjust_column_widths(ws)

        return wb

    def _setup_headers(self, ws, start_row: int = 3):
        # Structure: Employé | [Midi, H, Repas, Soir, H, Repas] x 7 jours | Total Heures, Total Repas
        # Les colonnes "H" (heures) seront cachées

        # Row 1: Main headers
        headers_row1 = ["Employé"]
        for day in DAY_LABELS_FR:
            headers_row1.extend([day, "", "", "", "", ""])  # 6 colonnes par jour
        headers_row1.extend(["Total Semaine", ""])

        # Row 2: Sub-headers
        headers_row2 = [""]
        for _ in DAY_LABELS_FR:
            headers_row2.extend(["Midi", "H", "Repas", "Soir", "H", "Repas"])
        headers_row2.extend(["Heures", "Repas"])

        # Write headers
        for col, header in enumerate(headers_row1, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        for col, header in enumerate(headers_row2, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Merge day headers (6 colonnes par jour maintenant)
        col = 2
        for _ in DAY_LABELS_FR:
            ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col + 5)
            col += 6

        # Merge weekly total header
        ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col + 1)

        # Cacher les colonnes d'heures (H) - colonnes 3, 6, 9, 12, ... (pattern: col 2 + 6*n + 1)
        for day_idx in range(7):
            # Colonne H midi = 2 + day_idx*6 + 1 = 3 + day_idx*6
            # Colonne H soir = 2 + day_idx*6 + 4 = 6 + day_idx*6
            h_midi_col = 3 + day_idx * 6
            h_soir_col = 6 + day_idx * 6
            ws.column_dimensions[get_column_letter(h_midi_col)].hidden = True
            ws.column_dimensions[get_column_letter(h_soir_col)].hidden = True

    def _populate_data(self, ws, planning: WeekPlanning, start_row: int = 5):
        for row_idx, employee in enumerate(planning.employees, start=start_row):
            # Employee name
            cell = ws.cell(row=row_idx, column=1, value=employee.name)
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

            col = 2
            meals_cols = []  # Pour stocker les colonnes des repas pour la formule
            hours_cols = []  # Pour stocker les colonnes des heures pour la formule

            for day in DAYS:
                day_schedule: DaySchedule = getattr(employee, day)

                # Afternoon time range
                cell = ws.cell(row=row_idx, column=col, value=day_schedule.afternoon.time_range or "-")
                cell.border = self.thin_border
                cell.alignment = self.center_align
                col += 1

                # Afternoon hours (colonne cachée)
                cell = ws.cell(row=row_idx, column=col, value=day_schedule.afternoon.hours if day_schedule.afternoon.hours else 0)
                cell.border = self.thin_border
                cell.alignment = self.center_align
                hours_cols.append(get_column_letter(col))
                col += 1

                # Afternoon meals
                afternoon_meals_val = day_schedule.afternoon.meals if day_schedule.afternoon.meals else 0
                cell = ws.cell(row=row_idx, column=col, value=afternoon_meals_val if afternoon_meals_val else "-")
                cell.border = self.thin_border
                cell.alignment = self.center_align
                meals_cols.append(get_column_letter(col))
                col += 1

                # Evening time range
                cell = ws.cell(row=row_idx, column=col, value=day_schedule.evening.time_range or "-")
                cell.border = self.thin_border
                cell.alignment = self.center_align
                col += 1

                # Evening hours (colonne cachée)
                cell = ws.cell(row=row_idx, column=col, value=day_schedule.evening.hours if day_schedule.evening.hours else 0)
                cell.border = self.thin_border
                cell.alignment = self.center_align
                hours_cols.append(get_column_letter(col))
                col += 1

                # Evening meals
                evening_meals_val = day_schedule.evening.meals if day_schedule.evening.meals else 0
                cell = ws.cell(row=row_idx, column=col, value=evening_meals_val if evening_meals_val else "-")
                cell.border = self.thin_border
                cell.alignment = self.center_align
                meals_cols.append(get_column_letter(col))
                col += 1

            # Weekly total hours - formule Excel pour additionner les colonnes d'heures cachées
            hours_formula = "=" + "+".join([f'{c}{row_idx}' for c in hours_cols])
            cell = ws.cell(row=row_idx, column=col, value=hours_formula)
            cell.border = self.thin_border
            cell.alignment = self.center_align
            cell.font = self.header_font
            col += 1

            # Weekly total meals - formule Excel pour additionner les repas (ignorer "-")
            meals_formula_parts = [f'IF({c}{row_idx}="-",0,{c}{row_idx})' for c in meals_cols]
            meals_formula = "=" + "+".join(meals_formula_parts)
            cell = ws.cell(row=row_idx, column=col, value=meals_formula)
            cell.border = self.thin_border
            cell.alignment = self.center_align
            cell.font = self.header_font

    def _adjust_column_widths(self, ws):
        ws.column_dimensions["A"].width = 18

        # 7 jours * 6 colonnes + 1 employé + 2 totaux = 45 colonnes
        # Pattern par jour: Midi(14), H(5), Repas(7), Soir(14), H(5), Repas(7)
        for day_idx in range(7):
            base_col = 2 + day_idx * 6
            ws.column_dimensions[get_column_letter(base_col)].width = 14      # Midi
            ws.column_dimensions[get_column_letter(base_col + 1)].width = 5   # H (caché)
            ws.column_dimensions[get_column_letter(base_col + 2)].width = 7   # Repas
            ws.column_dimensions[get_column_letter(base_col + 3)].width = 14  # Soir
            ws.column_dimensions[get_column_letter(base_col + 4)].width = 5   # H (caché)
            ws.column_dimensions[get_column_letter(base_col + 5)].width = 7   # Repas

        # Colonnes Total (après les 7 jours * 6 colonnes = 42 + 1 = colonne 44)
        total_col = 2 + 7 * 6  # = 44
        ws.column_dimensions[get_column_letter(total_col)].width = 8      # Total Heures
        ws.column_dimensions[get_column_letter(total_col + 1)].width = 7  # Total Repas

    def save_workbook(self, wb: Workbook, path: Path) -> Path:
        wb.save(path)
        return path

    def load_planning_from_excel(self, path: Path) -> WeekPlanning:
        wb = load_workbook(path)
        ws = wb.active

        # Extract week number from sheet title
        week_number = 1
        year = datetime.now().year

        if ws.title:
            if ws.title.startswith("Semaine "):
                try:
                    week_number = int(ws.title.split(" ")[1])
                except (ValueError, IndexError):
                    pass
            elif ws.title.startswith("Week "):
                try:
                    week_number = int(ws.title.split(" ")[1])
                except (ValueError, IndexError):
                    pass

        # Detect if there's a title row (new format starts data at row 5)
        # Old format starts at row 3
        first_cell = ws.cell(row=1, column=1).value
        if first_cell and "Planning des employés" in str(first_cell):
            data_start_row = 5
        else:
            data_start_row = 3

        employees = []
        row = data_start_row

        # Détecter si c'est l'ancien format (4 colonnes par jour) ou le nouveau (6 colonnes par jour)
        # En vérifiant le header de la ligne 4 ou 3+1
        header_row = data_start_row - 1
        # Vérifier si colonne 3 contient "H" (nouveau format) ou "Repas" (ancien format)
        col3_header = ws.cell(row=header_row, column=3).value or ""
        is_new_format = str(col3_header).strip() == "H"
        cols_per_day = 6 if is_new_format else 4

        while True:
            employee_name = ws.cell(row=row, column=1).value
            if not employee_name or str(employee_name).upper() == "TOTAL":
                break

            employee = EmployeeWeekSchedule(name=str(employee_name))
            col = 2

            for day in DAYS:
                if is_new_format:
                    # Nouveau format: Midi, H, Repas, Soir, H, Repas
                    afternoon_time = ws.cell(row=row, column=col).value or ""
                    afternoon_start, afternoon_end = self._parse_time_range(afternoon_time)
                    # col+1 = H (heures, ignoré car calculé)
                    afternoon_meals = ws.cell(row=row, column=col + 2).value
                    afternoon_meals = int(afternoon_meals) if afternoon_meals and afternoon_meals != "-" else 0

                    evening_time = ws.cell(row=row, column=col + 3).value or ""
                    evening_start, evening_end = self._parse_time_range(evening_time)
                    # col+4 = H (heures, ignoré car calculé)
                    evening_meals = ws.cell(row=row, column=col + 5).value
                    evening_meals = int(evening_meals) if evening_meals and evening_meals != "-" else 0
                else:
                    # Ancien format: Midi, Repas, Soir, Repas
                    afternoon_time = ws.cell(row=row, column=col).value or ""
                    afternoon_start, afternoon_end = self._parse_time_range(afternoon_time)
                    afternoon_meals = ws.cell(row=row, column=col + 1).value
                    afternoon_meals = int(afternoon_meals) if afternoon_meals and afternoon_meals != "-" else 0

                    evening_time = ws.cell(row=row, column=col + 2).value or ""
                    evening_start, evening_end = self._parse_time_range(evening_time)
                    evening_meals = ws.cell(row=row, column=col + 3).value
                    evening_meals = int(evening_meals) if evening_meals and evening_meals != "-" else 0

                day_schedule = DaySchedule(
                    afternoon=ShiftData(start_time=afternoon_start, end_time=afternoon_end, meals=afternoon_meals),
                    evening=ShiftData(start_time=evening_start, end_time=evening_end, meals=evening_meals),
                )
                setattr(employee, day, day_schedule)
                col += cols_per_day

            employees.append(employee)
            row += 1

        return WeekPlanning(
            week_number=week_number,
            year=year,
            employees=employees,
        )

    def _parse_time_range(self, time_str: str) -> tuple[str, str]:
        """Parse time range string like '11:30 - 14:30' into (start, end)."""
        if not time_str or time_str == "-":
            return "", ""

        match = re.match(r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})", str(time_str))
        if match:
            return match.group(1), match.group(2)
        return "", ""

    def update_planning_in_excel(self, path: Path, planning: WeekPlanning) -> Path:
        wb = self.create_planning_workbook(planning)
        self.save_workbook(wb, path)
        return path


excel_handler = ExcelHandler()
